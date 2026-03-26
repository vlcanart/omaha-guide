import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from collections import Counter

with open('omaha_master_enriched.json', encoding='utf-8') as f:
    events = json.load(f)

days_map = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
wb = Workbook()

# ═══════════════════════════════════════════════════════════
# SHEET 1: MASTER EVENTS
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Master Events'

# Title
ws.merge_cells('A1:S1')
t = ws['A1']
t.value = 'GO: Guide to Omaha - Master Event File v1.0 | March 23 - April 22, 2026'
t.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
t.fill = PatternFill('solid', fgColor='1B2A4A')
t.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# Subtitle
ws.merge_cells('A2:S2')
s = ws['A2']
s.value = f'{len(events)} events | 10+ sources | Generated {datetime.now().strftime("%Y-%m-%d")} | Enriched with YouTube, Spotify, artist bios, venue addresses'
s.font = Font(name='Arial', size=9, italic=True, color='666666')
s.alignment = Alignment(horizontal='center')

# Headers
headers = [
    'Date', 'Day', 'Time', 'Event', 'Category', 'Subcategory/Genre',
    'Venue', 'Venue Address', 'Area', 'Price', 'Age',
    'Artist Bio', 'Song 1 (YouTube)', 'Song 2 (YouTube)', 'Song 3 (YouTube)',
    'Best Video (YouTube)', 'Spotify', 'Ticket URL', 'Source'
]
hfill = PatternFill('solid', fgColor='2C5F8A')
hfont = Font(name='Arial', size=9, bold=True, color='FFFFFF')
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.font = hfont
    c.fill = hfill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[3].height = 32

# Category styling
cat_bg = {'concerts':'E8D5F5','sports':'D5ECD5','comedy':'FFF3D5','arts':'D5E8F5','family':'FFE0E0','festivals':'FFECD5'}
cat_fg = {'concerts':'7B2D8E','sports':'2D6B2D','comedy':'8B6914','arts':'2D5A8B','family':'8B2D2D','festivals':'8B5A14'}
bdr = Border(bottom=Side(style='thin', color='E0E0E0'))
link_font = Font(name='Arial', size=8, color='2255AA', underline='single')
alt = [PatternFill('solid', fgColor='FFFFFF'), PatternFill('solid', fgColor='F7F9FC')]

for i, ev in enumerate(events):
    r = i + 4
    fill = alt[i % 2]
    dt = datetime.strptime(ev['date'], '%Y-%m-%d')
    day = days_map[dt.weekday()]

    ws.cell(row=r, column=1, value=ev['date']).font = Font(name='Arial', size=9)
    ws.cell(row=r, column=2, value=day).font = Font(name='Arial', size=9)
    ws.cell(row=r, column=3, value=ev.get('time','TBD')).font = Font(name='Arial', size=9)
    ws.cell(row=r, column=4, value=ev['title']).font = Font(name='Arial', size=9, bold=True)

    cat = ev.get('cat','')
    cc = ws.cell(row=r, column=5, value=cat.title())
    cc.font = Font(name='Arial', size=8, bold=True, color=cat_fg.get(cat,'333333'))
    if cat in cat_bg:
        cc.fill = PatternFill('solid', fgColor=cat_bg[cat])
    cc.alignment = Alignment(horizontal='center')

    ws.cell(row=r, column=6, value=ev.get('subcategory','')).font = Font(name='Arial', size=8, color='555555')
    ws.cell(row=r, column=7, value=ev.get('venue','')).font = Font(name='Arial', size=9)
    ws.cell(row=r, column=8, value=ev.get('venueAddress','')).font = Font(name='Arial', size=8, color='666666')
    ws.cell(row=r, column=9, value=ev.get('area','Omaha')).font = Font(name='Arial', size=9)
    ws.cell(row=r, column=10, value=ev.get('price','TBD')).font = Font(name='Arial', size=9, bold=True)
    ws.cell(row=r, column=11, value=ev.get('ageRestriction','')).font = Font(name='Arial', size=8)
    ws.cell(row=r, column=12, value=ev.get('artistBio','')).font = Font(name='Arial', size=8, color='444444')

    # YouTube Song URLs
    for col_idx, field in [(13,'song1'), (14,'song2'), (15,'song3'), (16,'bestVideo')]:
        url = ev.get(field,'')
        cell = ws.cell(row=r, column=col_idx)
        if url:
            cell.value = url
            try:
                cell.hyperlink = url
                cell.font = link_font
            except:
                cell.font = Font(name='Arial', size=8, color='2255AA')

    # Spotify
    spot = ev.get('spotifyUrl','')
    sc = ws.cell(row=r, column=17)
    if spot:
        sc.value = spot
        try:
            sc.hyperlink = spot
            sc.font = link_font
        except:
            sc.font = Font(name='Arial', size=8, color='1DB954')

    # Ticket URL
    turl = ev.get('url','')
    tc = ws.cell(row=r, column=18)
    if turl:
        tc.value = turl
        try:
            tc.hyperlink = turl
            tc.font = link_font
        except:
            tc.font = Font(name='Arial', size=8, color='2255AA')

    ws.cell(row=r, column=19, value=ev.get('sourceId','')).font = Font(name='Arial', size=8, color='999999')

    # Apply alternating rows and borders
    for c in range(1, 20):
        cell = ws.cell(row=r, column=c)
        if c != 5:
            cell.fill = fill
        cell.border = bdr
        cell.alignment = Alignment(vertical='center', wrap_text=(c in [4, 8, 12]))

# Column widths
widths = {
    'A':11, 'B':5, 'C':9, 'D':40, 'E':10, 'F':22,
    'G':26, 'H':34, 'I':8, 'J':18, 'K':12,
    'L':50, 'M':28, 'N':28, 'O':28,
    'P':28, 'Q':28, 'R':50, 'S':16
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = 'A4'
ws.auto_filter.ref = f'A3:S{len(events)+3}'

# ═══════════════════════════════════════════════════════════
# SHEET 2: SUMMARY DASHBOARD
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Summary Dashboard')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'GO: Guide to Omaha - Content Summary'
ws2['A1'].font = Font(name='Arial', size=14, bold=True, color='1B2A4A')

# Category breakdown
ws2['A3'] = 'Category'; ws2['B3'] = 'Count'; ws2['C3'] = 'With YouTube'; ws2['D3'] = 'With Bio'
for c in ['A3','B3','C3','D3']:
    ws2[c].font = hfont; ws2[c].fill = hfill

cats = Counter(e['cat'] for e in events)
row = 4
for cat, cnt in cats.most_common():
    yt = sum(1 for e in events if e['cat'] == cat and e.get('song1'))
    bio = sum(1 for e in events if e['cat'] == cat and e.get('artistBio'))
    ws2.cell(row=row, column=1, value=cat.title())
    ws2.cell(row=row, column=2, value=cnt)
    ws2.cell(row=row, column=3, value=yt)
    ws2.cell(row=row, column=4, value=bio)
    row += 1
ws2.cell(row=row, column=1, value='TOTAL').font = Font(name='Arial', bold=True)
ws2.cell(row=row, column=2, value=len(events)).font = Font(name='Arial', bold=True)
ws2.cell(row=row, column=3, value=sum(1 for e in events if e.get('song1'))).font = Font(name='Arial', bold=True)
ws2.cell(row=row, column=4, value=sum(1 for e in events if e.get('artistBio'))).font = Font(name='Arial', bold=True)

# Subcategory breakdown
row += 2
ws2[f'A{row}'] = 'Subcategories / Genres'
ws2[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1B2A4A')
row += 1
ws2.cell(row=row, column=1, value='Subcategory').font = hfont; ws2.cell(row=row, column=1).fill = hfill
ws2.cell(row=row, column=2, value='Count').font = hfont; ws2.cell(row=row, column=2).fill = hfill
subs = Counter(e.get('subcategory','(none)') for e in events if e.get('subcategory'))
for sub, cnt in subs.most_common():
    row += 1
    ws2.cell(row=row, column=1, value=sub)
    ws2.cell(row=row, column=2, value=cnt)

# Top venues
row += 2
ws2[f'A{row}'] = 'Top Venues'
ws2[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1B2A4A')
row += 1
ws2.cell(row=row, column=1, value='Venue').font = hfont; ws2.cell(row=row, column=1).fill = hfill
ws2.cell(row=row, column=2, value='Events').font = hfont; ws2.cell(row=row, column=2).fill = hfill
ws2.cell(row=row, column=3, value='Address').font = hfont; ws2.cell(row=row, column=3).fill = hfill
venues = Counter(e.get('venue','') for e in events)
for v, c in venues.most_common(20):
    row += 1
    ws2.cell(row=row, column=1, value=v)
    ws2.cell(row=row, column=2, value=c)
    addr = next((e.get('venueAddress','') for e in events if e.get('venue') == v), '')
    ws2.cell(row=row, column=3, value=addr).font = Font(name='Arial', size=9, color='666666')

# Sources
row += 2
ws2[f'A{row}'] = 'Data Sources'
ws2[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1B2A4A')
row += 1
ws2.cell(row=row, column=1, value='Source').font = hfont; ws2.cell(row=row, column=1).fill = hfill
ws2.cell(row=row, column=2, value='Events').font = hfont; ws2.cell(row=row, column=2).fill = hfill
for src, c in Counter(e.get('sourceId','') for e in events).most_common():
    row += 1
    ws2.cell(row=row, column=1, value=src)
    ws2.cell(row=row, column=2, value=c)

# Enrichment stats
row += 2
ws2[f'A{row}'] = 'Enrichment Coverage'
ws2[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1B2A4A')
row += 1
stats = [
    ('Subcategory/Genre', sum(1 for e in events if e.get('subcategory'))),
    ('Artist Bio', sum(1 for e in events if e.get('artistBio'))),
    ('YouTube Songs (1+)', sum(1 for e in events if e.get('song1'))),
    ('Best YouTube Video', sum(1 for e in events if e.get('bestVideo'))),
    ('Spotify URL', sum(1 for e in events if e.get('spotifyUrl'))),
    ('Venue Address', sum(1 for e in events if e.get('venueAddress'))),
    ('Age Restriction', sum(1 for e in events if e.get('ageRestriction'))),
    ('Ticket URL', sum(1 for e in events if e.get('url'))),
]
ws2.cell(row=row, column=1, value='Field').font = hfont; ws2.cell(row=row, column=1).fill = hfill
ws2.cell(row=row, column=2, value='Populated').font = hfont; ws2.cell(row=row, column=2).fill = hfill
ws2.cell(row=row, column=3, value='Coverage').font = hfont; ws2.cell(row=row, column=3).fill = hfill
for label, count in stats:
    row += 1
    ws2.cell(row=row, column=1, value=label)
    ws2.cell(row=row, column=2, value=count)
    pct = f'{count/len(events)*100:.0f}%'
    ws2.cell(row=row, column=3, value=pct).font = Font(name='Arial', bold=True,
        color='2D6B2D' if count/len(events) > 0.8 else '8B6914' if count/len(events) > 0.3 else '8B2D2D')

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 36
ws2.column_dimensions['D'].width = 12

# Save
out = 'omaha-events-master-v1.xlsx'
wb.save(out)
print(f'Saved {out} with {len(events)} events, 19 columns, 2 sheets')
